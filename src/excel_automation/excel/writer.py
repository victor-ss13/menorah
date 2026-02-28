"""Geração de arquivos Excel formatados com openpyxl."""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Cor do cabeçalho (azul escuro)
HEADER_FILL_COLOR = "1F3864"
HEADER_FONT_COLOR = "FFFFFF"


class ExcelWriter:
    """Gera planilhas Excel formatadas a partir de um :class:`pandas.DataFrame`."""

    def write(
        self,
        df: pd.DataFrame,
        output_path: Path | str,
        sheet_name: str = "Dados",
    ) -> Path:
        """Escreve o DataFrame em um arquivo Excel formatado.

        O arquivo gerado possui:
        - Cabeçalho com fundo azul escuro e texto branco em negrito.
        - Colunas com largura ajustada automaticamente ao conteúdo.
        - Linhas alternadas com fundo cinza claro para facilitar leitura.
        - Filtros automáticos no cabeçalho.

        Args:
            df: DataFrame a ser escrito.
            output_path: Caminho do arquivo de saída (``.xlsx``).
            sheet_name: Nome da aba (padrão: ``"Dados"``).

        Returns:
            :class:`pathlib.Path` do arquivo gerado.

        Raises:
            OSError: Se não for possível criar o arquivo.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        # --- Cabeçalho ---
        header_fill = PatternFill(
            fill_type="solid", fgColor=HEADER_FILL_COLOR
        )
        header_font = Font(bold=True, color=HEADER_FONT_COLOR)

        for col_idx, column_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=str(column_name))
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Dados ---
        alt_fill = PatternFill(fill_type="solid", fgColor="E8EAED")
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill

        # --- Auto-fit de largura de colunas ---
        for col_idx, column_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(column_name)),
                df.iloc[:, col_idx - 1].astype(str).str.len().max()
                if len(df) > 0
                else 0,
            )
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

        # --- Filtros automáticos ---
        ws.auto_filter.ref = ws.dimensions

        wb.save(output_path)
        logger.info("Planilha salva em: %s (%d linhas)", output_path, len(df))
        return output_path
